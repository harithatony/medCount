import cv2
from list3 import YOLO_Pred
import pyttsx3
class Detection:
    def __init__(self, model_path, config_path):
        self.yolo = YOLO_Pred(model_path, config_path)

    def counting(self, img_path):
        img = cv2.imread(img_path)

        # Predictions
        result = self.yolo.predictions(img)
        if isinstance(result, tuple) and len(result) == 2:
            img_pred, before = result
        else:
            img_pred = result
            before = []

        # Making dictionary of before
        b = {}
        for i in before:
            if i not in b:
                b[i] = 1
            else:
                b[i] += 1

        return img_pred, b

# Example usage
def count_instruments(image_path):
    detection = Detection('./Model/weight/best.onnx', 'data_yaml')
    return detection.counting(image_path)

def check(before,after):
    b={}
    for i in before:
        for j in i:
            if j in b:
                b[j] += i[j]
            else:
                b[j] = i[j]
    # print(b)
    a={}
    for i in after:
        for j in i:
            if j in a:
                a[j] += i[j]
            else:
                a[j] = i[j]
    diff = {}
    for i in b:
        if i not in a :
            diff[i] = b[i]
        elif b[i]  != a[i]:
            diff[i] = abs(b[i]-a[i])
    for i in a:
        if i not in b :
            diff[i] = a[i]
    # print(diff,"hasdfghj")
    return diff
before = [{'Curved Mayo Scissor': 4, 'Straight Mayo Scissor': 2, 'Straight Dissection Clamp': 1}]
after = []
check(before,after)


def speak_warning_message(msg):
       # Initialize the Text-to-Speech engine
        engine = pyttsx3.init()

        # Set properties (optional)
        engine.setProperty('rate', 150)  # Speed of speech
        engine.setProperty('volume', 0.9)  # Volume level (0.0 to 1.0)

        # Warning message
        warning_message = msg

        # Speak the warning message
        engine.say(warning_message)

        # Wait for speech to finish
        engine.runAndWait()
 
import openpyxl

def add_patient_to_excel(patient_name, doctor_name,staff_name,Status):
    # Open the Excel file
    wb = openpyxl.load_workbook("patient_data.xlsx")
    sheet = wb.active
    # for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):  # start from 2nd row, 1st column
    #     if row[0] == patient_name:  # Check if value in the first column is "soos"
    #         # Update the value in the 4th column (D)
    #         sheet.cell(row=row[0], column=4, value=None)
    #         sheet.cell(row=row[0], column=4, value=Status)
    #         break
    # else:
    # Find the next available row
    next_row = sheet.max_row + 1

        # Add patient name and doctor name to the Excel sheet
    sheet.cell(row=next_row, column=1, value=patient_name)
    sheet.cell(row=next_row, column=2, value=doctor_name)
    sheet.cell(row=next_row, column=3, value=staff_name)
    sheet.cell(row=next_row, column=4, value=Status)
    # Save the changes
    wb.save("patient_data.xlsx")


# Example usage:

